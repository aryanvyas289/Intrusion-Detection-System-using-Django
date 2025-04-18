import os
import datetime
from django.core.management.base import BaseCommand
from django.utils.timezone import make_aware, localtime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from IDS.models import BruteForceDetection, SQLInjectionDetection, DOSDetection

class Command(BaseCommand):
    help = "Generate an Excel report of attacks on a specific date for a given attack type or all attack types."

    ATTACK_MODELS = {
        "BruteForce": BruteForceDetection,
        "SQLInjection": SQLInjectionDetection,
        "DOS": DOSDetection,
    }

    def add_arguments(self, parser):
        parser.add_argument(
            '--attack',
            type=str,
            help='Type of attack (BruteForce, SQLInjection, DOS) or "all" for all attack types'
        )
        parser.add_argument(
            '--date',
            type=str,
            required=True,
            help='Date in YYYY-MM-DD format'
        )

    def handle(self, *args, **kwargs):
        attack_type = kwargs.get('attack')
        date_str = kwargs['date']

        # Validate date
        try:
            report_date = make_aware(datetime.datetime.strptime(date_str, "%Y-%m-%d"))
        except ValueError:
            self.stderr.write(self.style.ERROR("Invalid date format. Use YYYY-MM-DD."))
            return

        start_of_day = report_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = report_date.replace(hour=23, minute=59, second=59, microsecond=999999)

        # Initialize workbook
        wb = Workbook()
        # Remove default sheet created by Workbook()
        wb.remove(wb.active)

        # Define styles
        header_font = Font(name="Verdana", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        data_font = Font(name="Verdana", size=10, bold=False)

        # Handle specific attack type or all
        attack_types = [attack_type] if attack_type and attack_type != "all" else self.ATTACK_MODELS.keys()

        if attack_type and attack_type != "all" and attack_type not in self.ATTACK_MODELS:
            self.stderr.write(self.style.ERROR(f"Invalid attack type. Choose from {list(self.ATTACK_MODELS.keys())} or 'all'"))
            return

        report_generated = False

        for atk_type in attack_types:
            model = self.ATTACK_MODELS[atk_type]
            attacks = model.objects.filter(Detection_date_and_time__range=(start_of_day, end_of_day))

            if not attacks.exists():
                self.stdout.write(self.style.WARNING(f"No {atk_type} attack records found for {date_str}."))
                continue

            report_generated = True
            ws = wb.create_sheet(title=f"{atk_type}_Attacks_{date_str}")

            # Define headers
            if atk_type == "BruteForce":
                headers = ["Date", "Time", "Attacker's IP", "Username", "Password", "Attempts"]
            elif atk_type == "SQLInjection":
                headers = ["Date", "Time", "Attacker's IP", "Username", "Password"]
            else:  # DOS
                headers = ["Date", "Time", "Attacker's IP", "Attack Type", "Traffic Rate", "Details"]
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            ws.row_dimensions[1].height = 25

            # Add data rows
            for attack in attacks:
                local_time = localtime(attack.Detection_date_and_time)
                if atk_type == "BruteForce":
                    ws.append([
                        local_time.strftime('%Y-%m-%d'),
                        local_time.strftime('%H:%M:%S'),
                        attack.Attackers_IP,
                        attack.attempted_username,
                        attack.attempted_password,
                        attack.Number_of_attempts
                    ])
                elif atk_type == "SQLInjection":
                    ws.append([
                        local_time.strftime('%Y-%m-%d'),
                        local_time.strftime('%H:%M:%S'),
                        attack.Attackers_IP,
                        attack.attempted_username,
                        attack.attempted_password
                    ])
                else:  # DOS
                    ws.append([
                        local_time.strftime('%Y-%m-%d'),
                        local_time.strftime('%H:%M:%S'),
                        attack.Attackers_IP or "N/A",
                        attack.Attack_type or "N/A",
                        attack.Traffic_rate if attack.Traffic_rate is not None else "N/A",
                        attack.Details or "N/A"
                    ])

            # Apply font style to data rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.font = data_font

            # Adjust column widths
            ws.column_dimensions['A'].width = 15  # Date
            ws.column_dimensions['B'].width = 12  # Time
            ws.column_dimensions['C'].width = 18  # Attacker's IP
            ws.column_dimensions['D'].width = 18  # Username or Attack Type
            ws.column_dimensions['E'].width = 18  # Password or Traffic Rate
            if atk_type == "BruteForce":
                ws.column_dimensions['F'].width = 15  # Attempts
            elif atk_type == "DOS":
                ws.column_dimensions['F'].width = 30  # Details

        if not report_generated:
            self.stdout.write(self.style.WARNING(f"No attack records found for any attack type on {date_str}."))
            return

        # Save file
        filename = f"{'all_attacks' if attack_type == 'all' else attack_type.lower()}_report_{date_str}.xlsx"
        report_path = os.path.join(os.getcwd(), filename)
        wb.save(report_path)

        self.stdout.write(self.style.SUCCESS(f"Report saved: {report_path}"))